"""
buscador_dados.py — Coleta de dados ANBIMA via pyettj e histórico XLSX
"""
import os
import pandas as pd
from datetime import date, timedelta
import logging

logger = logging.getLogger(__name__)

# ── Vértices padrão DU ───────────────────────────────────────────────────
# 13 pontos publicados pela ANBIMA (seções PREFIXADOS + ETTJ PREF).
# DU 84 e 189 não são publicados; a interpolação flat forward os cobre.
VERTICES_PADRAO_DU = [
    21, 42, 63, 126, 252, 378,
    504, 630, 756, 882, 1008, 1134, 1260,
]

# 9 vértices "de mercado" usados na aba Movimentos (comparação hoje vs 30d)
VERTICES_MOVIMENTOS = {21, 42, 63, 126, 252, 504, 756, 1008, 1260}

# ── Dados estáticos de fallback ────────────────────────────────────────────
FALLBACK_VERTICES = [
    (21,   0.1495), (42,  0.1492), (63,  0.1489),
    (126,  0.1478), (252, 0.1462), (378, 0.1454),
    (504,  0.1446), (630, 0.1440), (756, 0.1435), (882, 0.1431),
    (1008, 0.1427), (1134, 0.1426), (1260, 0.1425),
]
FALLBACK_VERTICES_30D = [
    (21,   0.1460), (42,  0.1458), (63,  0.1455),
    (126,  0.1442), (252, 0.1430), (378, 0.1422),
    (504,  0.1415), (630, 0.1410), (756, 0.1405), (882, 0.1401),
    (1008, 0.1398), (1134, 0.1396), (1260, 0.1395),
]

NOME_SHEET    = "Historico"
COLUNAS_XLSX  = ["Data Referencia", "Prazo (DU)", "Taxa Spot (% a.a.)"]


# ── Helpers ────────────────────────────────────────────────────────────────

def _data_str(d: date) -> str:
    return d.strftime("%d/%m/%Y")


def _dia_util_anterior(d: date, n: int) -> date:
    """Retorna o dia útil n dias úteis atrás usando calendário ANBIMA (bizdays)."""
    try:
        import bizdays
        cal = bizdays.Calendar.load("ANBIMA")
        result = cal.offset(d, -n)
        return result.date() if hasattr(result, "date") else result
    except Exception:
        logger.warning("bizdays indisponível — usando contagem simples de dias úteis.")
        current = d
        count = 0
        while count < n:
            current -= timedelta(days=1)
            if current.weekday() < 5:
                count += 1
        return current


# ── Histórico XLSX ─────────────────────────────────────────────────────────

def _ler_historico_xlsx(historico_path: str) -> pd.DataFrame | None:
    """Lê o histórico xlsx. Retorna DataFrame com [data_referencia, dias_uteis, taxa] ou None."""
    if not os.path.exists(historico_path):
        return None
    try:
        df = pd.read_excel(historico_path, sheet_name=NOME_SHEET, engine="openpyxl")
        df.columns = ["data_referencia", "dias_uteis", "taxa"]
        df["data_referencia"] = pd.to_datetime(df["data_referencia"]).dt.date
        df["dias_uteis"] = pd.to_numeric(df["dias_uteis"], errors="coerce").astype("Int64")
        df["taxa"]       = pd.to_numeric(df["taxa"], errors="coerce")
        return df.dropna().reset_index(drop=True)
    except Exception as e:
        logger.warning(f"Erro lendo histórico xlsx: {e}")
        return None


def _escrever_historico_xlsx(df_total: pd.DataFrame, historico_path: str):
    """Reescreve o histórico xlsx completo com formatação dark navy."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    os.makedirs(os.path.dirname(historico_path), exist_ok=True)

    COR_HEADER  = "243356"
    CORES_DADOS = ["162240", "1C2B4A"]   # alternância por data

    borda_lado = Side(style="thin", color="2A3F6A")
    borda  = Border(left=borda_lado, right=borda_lado,
                    top=borda_lado, bottom=borda_lado)
    centro = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill("solid", fgColor=COR_HEADER)
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    data_font   = Font(color="FFFFFF", name="Calibri", size=10)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = NOME_SHEET

    # Cabeçalhos
    for col, nome in enumerate(COLUNAS_XLSX, 1):
        c = ws.cell(row=1, column=col, value=nome)
        c.fill      = header_fill
        c.font      = header_font
        c.alignment = centro
        c.border    = borda

    # Índice de cor por data (alternância)
    datas_unicas = sorted(df_total["data_referencia"].unique())
    cor_por_data = {d: i % 2 for i, d in enumerate(datas_unicas)}

    # Dados
    for row_idx, row in enumerate(df_total.itertuples(index=False), start=2):
        fill = PatternFill("solid", fgColor=CORES_DADOS[cor_por_data[row.data_referencia]])

        c1 = ws.cell(row=row_idx, column=1, value=str(row.data_referencia))
        c2 = ws.cell(row=row_idx, column=2, value=int(row.dias_uteis))
        c3 = ws.cell(row=row_idx, column=3, value=float(row.taxa))
        c3.number_format = "0.00%"

        for c in [c1, c2, c3]:
            c.fill      = fill
            c.font      = data_font
            c.alignment = centro
            c.border    = borda

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 22
    ws.freeze_panes = "A2"
    ws.sheet_properties.tabColor = "243356"

    wb.save(historico_path)


def _migrar_csv_para_xlsx(csv_path: str, xlsx_path: str):
    """Migra o CSV legado para xlsx formatado."""
    try:
        df = pd.read_csv(csv_path)
        df.columns = ["data_referencia", "dias_uteis", "taxa"]
        df["data_referencia"] = pd.to_datetime(df["data_referencia"]).dt.date
        df = df.sort_values(["data_referencia", "dias_uteis"]).reset_index(drop=True)
        _escrever_historico_xlsx(df, xlsx_path)
        logger.info(f"CSV migrado para xlsx: {len(df)} registros. "
                    f"O arquivo '{csv_path}' pode ser deletado manualmente.")
    except Exception as e:
        logger.warning(f"Migração CSV→xlsx falhou: {e}")


def salvar_historico(data: date, curva: pd.DataFrame, historico_path: str):
    """Salva curva no histórico xlsx (sem duplicar). Migra CSV legado se necessário."""
    # Migrar CSV legado se xlsx ainda não existe
    if not os.path.exists(historico_path):
        csv_path = historico_path.replace(".xlsx", ".csv")
        if os.path.exists(csv_path):
            _migrar_csv_para_xlsx(csv_path, historico_path)

    # Verificar duplicata — sobrescrever se dados novos tiverem mais vértices
    df_existente = _ler_historico_xlsx(historico_path)
    if df_existente is not None and data in df_existente["data_referencia"].values:
        n_existente = len(df_existente[df_existente["data_referencia"] == data])
        if len(curva) <= n_existente:
            return
        df_existente = df_existente[df_existente["data_referencia"] != data].reset_index(drop=True)

    # Montar DataFrame total e reescrever
    df_novo = curva[["dias_uteis", "taxa"]].copy()
    df_novo.insert(0, "data_referencia", data)

    if df_existente is not None:
        df_total = pd.concat([df_existente, df_novo], ignore_index=True)
    else:
        df_total = df_novo

    df_total = df_total.sort_values(["data_referencia", "dias_uteis"]).reset_index(drop=True)
    _escrever_historico_xlsx(df_total, historico_path)
    logger.info(f"Histórico xlsx atualizado: {data} ({len(curva)} vértices)")


def _buscar_historico(data: date, historico_path: str) -> pd.DataFrame | None:
    """Busca curva no histórico local. Retorna None se entrada tiver poucos vértices."""
    df = _ler_historico_xlsx(historico_path)
    if df is None:
        return None
    subset = df[df["data_referencia"] == data][["dias_uteis", "taxa"]].reset_index(drop=True)
    min_vertices = max(5, len(VERTICES_PADRAO_DU) - 2)
    if len(subset) >= min_vertices:
        logger.info(f"[historico] {data}: {len(subset)} vértices")
        return subset
    if len(subset) > 0:
        logger.info(f"[historico] {data}: {len(subset)} vértices (insuficiente, re-buscando)")
    return None


def _buscar_data_mais_proxima(data_alvo: date, historico_path: str) -> tuple[pd.DataFrame | None, date | None]:
    """Busca a data mais próxima (antes ou depois) disponível no histórico."""
    df = _ler_historico_xlsx(historico_path)
    if df is None:
        return None, None
    datas = sorted(df["data_referencia"].unique())
    if not datas:
        return None, None
    data_escolhida = min(datas, key=lambda d: abs((d - data_alvo).days))
    subset = df[df["data_referencia"] == data_escolhida][["dias_uteis", "taxa"]].reset_index(drop=True)
    if len(subset) >= 5:
        return subset, data_escolhida
    return None, None


def _buscar_anbima_csv(data: date | None = None) -> tuple[pd.DataFrame | None, date | None]:
    """
    Busca curva prefixada direto do endpoint CSV da ANBIMA.
    Retorna (DataFrame com [dias_uteis, taxa], data_referencia) ou (None, None).

    Combina duas seções do CSV:
      - 'ETTJ PREF' (coluna da seção de inflação implícita): 126, 378, 630, …
      - 'PREFIXADOS (CIRCULAR 3.361)': vértices curtos (21, 42, 63, …)

    Se ``data`` for informada, passa Dt_Ref=dd/mm/yyyy para obter dados históricos.
    Sem ``data``, retorna o dia mais recente publicado pela ANBIMA.
    """
    import re
    import requests
    from datetime import datetime as _dt

    URL = "https://www.anbima.com.br/informacoes/est-termo/CZ-down.asp"
    params = {"Dt_Ref": _data_str(data)} if data else {}
    try:
        r = requests.get(URL, params=params, timeout=15)
        r.raise_for_status()
        texto = r.text
    except Exception as e:
        logger.warning(f"[anbima_csv] Falha na requisição: {e}")
        return None, None

    if not texto.strip():
        logger.warning(f"[anbima_csv] Resposta vazia para {data}")
        return None, None

    # ── Data de referência ────────────────────────────────────────────────
    data_ref = None
    primeira_linha = texto.splitlines()[0]
    match = re.match(r"(\d{2}/\d{2}/\d{4})", primeira_linha)
    if match:
        try:
            data_ref = _dt.strptime(match.group(1), "%d/%m/%Y").date()
        except Exception:
            pass

    def _parse_br(s: str) -> float:
        return float(s.strip().replace(".", "").replace(",", "."))

    vertices: dict[int, float] = {}

    # ── Seção 1: ETTJ PREF (vértices intermediários: 378, 630, 882, 1134…) ──
    marcador_ettj = "ETTJ Infla"
    marcador_pref = "PREFIXADOS (CIRCULAR 3.361)"
    if marcador_ettj in texto and marcador_pref in texto:
        bloco_ettj = texto.split(marcador_ettj)[1].split(marcador_pref)[0]
        linhas_ettj = [l.strip() for l in bloco_ettj.splitlines() if l.strip()]
        for linha in linhas_ettj[1:]:
            partes = linha.split(";")
            if len(partes) < 3:
                continue
            try:
                du   = int(partes[0].replace(".", ""))
                taxa = _parse_br(partes[2]) / 100.0
                if taxa > 0:
                    vertices[du] = taxa
            except (ValueError, IndexError):
                continue

    # ── Seção 2: PREFIXADOS (CIRCULAR 3.361) — curto prazo (21, 42, 63…) ──
    if marcador_pref in texto:
        bloco_pref = texto.split(marcador_pref)[1]
        linhas_pref = [l.strip() for l in bloco_pref.splitlines() if l.strip()]
        for linha in linhas_pref[1:]:
            partes = linha.split(";")
            if len(partes) < 2:
                break
            try:
                du   = int(partes[0].replace(".", ""))
                taxa = _parse_br(partes[1]) / 100.0
                vertices[du] = taxa
            except (ValueError, IndexError):
                break

    if len(vertices) < 5:
        logger.warning(f"[anbima_csv] Vértices insuficientes: {len(vertices)}")
        return None, None

    result = (
        pd.DataFrame(list(vertices.items()), columns=["dias_uteis", "taxa"])
        .sort_values("dias_uteis")
        .reset_index(drop=True)
    )
    logger.info(f"[anbima_csv] {data_ref}: {len(result)} vértices")
    return result, data_ref


def _buscar_pyettj(data: date) -> pd.DataFrame | None:
    """Busca via pyettj — extrai taxas nos vértices padrão DU."""
    try:
        from pyettj import ettj
        df_raw = ettj.get_ettj(_data_str(data), curva="PRE")
        if df_raw is None or len(df_raw) == 0:
            return None
        col_dc   = df_raw.columns[0]
        col_taxa = df_raw.columns[1]
        df_raw = df_raw[[col_dc, col_taxa]].dropna().copy()
        df_raw.columns = ["dc", "taxa_pct"]
        df_raw["dc"]      = pd.to_numeric(df_raw["dc"],      errors="coerce")
        df_raw["taxa_pct"] = pd.to_numeric(df_raw["taxa_pct"], errors="coerce")
        df_raw = df_raw.dropna()
        if len(df_raw) < len(VERTICES_PADRAO_DU):
            return None
        # Calcular DC correto para cada DU usando calendário ANBIMA (bizdays)
        # Fallback: aproximação 365/252 se bizdays indisponível
        try:
            import bizdays
            cal = bizdays.Calendar.load("ANBIMA")
            def _du_para_dc(ref: date, du: int) -> int:
                venc = cal.offset(ref, du)
                venc = venc.date() if hasattr(venc, "date") else venc
                return (venc - ref).days
        except Exception:
            logger.warning("bizdays indisponível — usando 365/252 para DC.")
            def _du_para_dc(ref: date, du: int) -> int:
                return round(du * 365 / 252)

        rows = []
        for du in VERTICES_PADRAO_DU:
            dc_alvo = _du_para_dc(data, du)
            idx  = (df_raw["dc"] - dc_alvo).abs().idxmin()
            taxa = float(df_raw.loc[idx, "taxa_pct"]) / 100.0
            rows.append({"dias_uteis": du, "taxa": taxa})
        result = pd.DataFrame(rows)
        logger.info(f"[pyettj] {data}: {len(result)} vértices")
        return result
    except Exception as e:
        logger.warning(f"pyettj falhou para {data}: {e}")
        return None



def _fallback_estatico(tipo: str = "hoje") -> pd.DataFrame:
    """Dados estáticos como último recurso."""
    logger.warning("Usando dados estáticos de fallback.")
    vertices = FALLBACK_VERTICES if tipo == "hoje" else FALLBACK_VERTICES_30D
    return pd.DataFrame(vertices, columns=["dias_uteis", "taxa"])


# ── Funções públicas ───────────────────────────────────────────────────────

def buscar_curva(data: date, historico_path: str) -> pd.DataFrame:
    """
    Busca curva ETTJ.
    Ordem: ANBIMA CSV direto → pyettj → histórico local → fallback estático.
    """
    # 1. ANBIMA CSV (endpoint atualizado — substitui pyettj como fonte primária)
    df, data_csv = _buscar_anbima_csv()
    if df is not None:
        data_efetiva = data_csv if data_csv is not None else data
        salvar_historico(data_efetiva, df, historico_path)
        if data_csv and data_csv < data:
            logger.warning(
                f"ANBIMA ainda não publicou dados para {data.strftime('%d/%m/%Y')} — "
                f"usando dados disponíveis de {data_csv.strftime('%d/%m/%Y')}."
            )
        return df

    # 2. pyettj (fallback caso o endpoint CSV mude novamente)
    df = _buscar_pyettj(data)
    if df is not None:
        salvar_historico(data, df, historico_path)
        return df
    data_anterior = _dia_util_anterior(data, 1)
    df = _buscar_pyettj(data_anterior)
    if df is not None:
        salvar_historico(data_anterior, df, historico_path)
        return df

    # 3. Histórico local
    df = _buscar_historico(data, historico_path)
    if df is not None:
        return df

    # 4. Último recurso
    return _fallback_estatico("hoje")


def buscar_curva_comparacao(data_ref: date, dias_atras: int, historico_path: str) -> tuple[pd.DataFrame, date]:
    """Busca curva de comparação (N dias úteis atrás). Retorna (curva, data_efetiva)."""
    data_alvo = _dia_util_anterior(data_ref, dias_atras)
    logger.info(f"Data comparação: {data_alvo}")

    # 1. ANBIMA CSV com Dt_Ref (fonte primária — busca histórica direta)
    df, data_csv = _buscar_anbima_csv(data_alvo)
    if df is not None:
        data_efetiva = data_csv if data_csv is not None else data_alvo
        salvar_historico(data_efetiva, df, historico_path)
        return df, data_efetiva

    # 2. Histórico local (exato)
    df = _buscar_historico(data_alvo, historico_path)
    if df is not None:
        return df, data_alvo

    # 3. pyettj (B3) como fallback
    df = _buscar_pyettj(data_alvo)
    if df is not None:
        salvar_historico(data_alvo, df, historico_path)
        return df, data_alvo

    # 4. Data mais próxima no histórico
    df, data_real = _buscar_data_mais_proxima(data_alvo, historico_path)
    if df is not None:
        logger.info(f"Usando data mais próxima: {data_real}")
        return df, data_real

    return _fallback_estatico("30d"), data_alvo
